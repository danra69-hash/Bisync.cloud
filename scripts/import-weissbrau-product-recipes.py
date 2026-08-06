#!/usr/bin/env python3
"""Import Weissbrau finished-product recipes from Product Component Excel.

Creates/updates B2C products for company 5 and attaches BOM lines to matching
Ingredients (components). Sub-product recipe lines are skipped — those products
are imported separately.

Source: data/weissbrau-product-recipe.xlsx
"""

from __future__ import annotations

import csv
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "weissbrau-product-recipe.xlsx"
MENU_CSV = ROOT / "data" / "weissbrau-menu-b2c.csv"
API = "https://bisync-cloud-389272498937.asia-southeast1.run.app"
COMPANY_ID = 5
LOCATION_IDS = ["weissbrau-pavilion-kuala-lumpur"]

UOM_MAP = {
    "MG": "mg",
    "GR": "g",
    "G": "g",
    "GRAM": "g",
    "GRAMS": "g",
    "KG": "kg",
    "ML": "ml",
    "CL": "cl",
    "LTR": "L",
    "L": "L",
    "LITRE": "L",
    "LITER": "L",
    "EACH": "pcs",
    "PCS": "pcs",
    "PC": "pcs",
    "BTL": "btl",
    "BOTTLE": "btl",
    "CAN": "can",
    "TIN": "tin",
    "SLICE": "slice",
    "PACK": "pack",
    "PKT": "pkt",
    "BAG": "bag",
    "BOX": "box",
    "CTN": "ctn",
    "CASE": "case",
    "PTN": "ptn",
    "PORTION": "ptn",
}

# Recipe ingredient name → existing component name (normalized match key uses these).
INGREDIENT_ALIASES = {
    "PUCK SHREDDED MOZZARELLA": "SHREDDED MOZZARELLA",
    "ICE CREAM VANILLA": "GELATO VANILLA",
    "GELATO COOKIE": "GELATO VANILLA",  # closest available gelato; note in report
    "ROUGH SALT": "SALT ROCK",
    "CUCUMBER FRESH LOCAL": "CUCUMBER JAPANESE KYURI",
}

# Missing wines / niche items — create as components if still absent.
STUB_COMPONENTS = {
    "BARBARESCO DOCG PIO CESARE": ("Beverage", "Red Wine", "ml", "btl"),
    "BARBERA D'ALBA DOCG": ("Beverage", "Red Wine", "ml", "btl"),
    "MIRABELLEN": ("Beverage", "German Schnapps", "ml", "btl"),
    "GELATO COOKIE": ("Food", "Desserts", "g", "kg"),
    "CUCUMBER FRESH LOCAL": ("Food", "Produce", "g", "kg"),
}

MENU_GROUP_CATEGORY = {
    "Starters": "Food",
    "Soups": "Food",
    "Salads": "Food",
    "Salad Add-On": "Food",
    "Good To Share": "Food",
    "Pasta": "Food",
    "Burgers & Sandwiches": "Food",
    "Schnitzel": "Food",
    "German Sausages": "Food",
    "Sausage Add-On": "Food",
    "Specialties": "Food",
    "Desserts": "Food",
    "Sauce & Dressing": "Food",
    "Draught Beer": "Beverage",
    "Bottled Beer": "Beverage",
    "Whisky": "Beverage",
    "Vodka": "Beverage",
    "Gin": "Beverage",
    "Rum": "Beverage",
    "Brandy": "Beverage",
    "Tequila": "Beverage",
    "Liqueur": "Beverage",
    "German Schnapps": "Beverage",
    "Red Wine": "Beverage",
    "White Wine": "Beverage",
    "Sparkling Wine": "Beverage",
    "Cocktails": "Beverage",
    "Mocktails": "Beverage",
    "Coffee": "Beverage",
    "Tea": "Beverage",
    "Fruit Juice & Ice Teas": "Beverage",
    "Soft Drinks": "Beverage",
    "Mineral Water": "Beverage",
    "Bev Promotion": "Beverage",
}

BEVERAGE_GROUP_KEYWORDS = [
    ("DRAUGHT", "Draught Beer"),
    ("PINT", "Draught Beer"),
    ("WEISSBIER", "Draught Beer"),
    ("PILSNER", "Draught Beer"),
    ("LAGER", "Draught Beer"),
    ("CIDER", "Bottled Beer"),
    ("BEER", "Bottled Beer"),
    ("WHISKY", "Whisky"),
    ("WHISKEY", "Whisky"),
    ("ABERLOUR", "Whisky"),
    ("GLEN", "Whisky"),
    ("MACALLAN", "Whisky"),
    ("CHIVAS", "Whisky"),
    ("ROYAL SALUTE", "Whisky"),
    ("JACK DANIEL", "Whisky"),
    ("JAMESON", "Whisky"),
    ("VODKA", "Vodka"),
    ("ABSOLUT", "Vodka"),
    ("GREY GOOSE", "Vodka"),
    ("BELVEDERE", "Vodka"),
    ("GIN", "Gin"),
    ("BOMBAY", "Gin"),
    ("TANQUERAY", "Gin"),
    ("HENDRICK", "Gin"),
    ("RUM", "Rum"),
    ("BACARDI", "Rum"),
    ("HAVANA", "Rum"),
    ("BRANDY", "Brandy"),
    ("COGNAC", "Brandy"),
    ("HENNESSY", "Brandy"),
    ("TEQUILA", "Tequila"),
    ("PATRON", "Tequila"),
    ("JOSE CUERVO", "Tequila"),
    ("LIQUEUR", "Liqueur"),
    ("BAILEYS", "Liqueur"),
    ("KAHLUA", "Liqueur"),
    ("SCHNAPS", "German Schnapps"),
    ("SCHNAPPS", "German Schnapps"),
    ("MIRABELLEN", "German Schnapps"),
    ("PROSECCO", "Sparkling Wine"),
    ("CHAMPAGNE", "Sparkling Wine"),
    ("SPARKLING", "Sparkling Wine"),
    ("BAROLO", "Red Wine"),
    ("BARBARESCO", "Red Wine"),
    ("BARBERA", "Red Wine"),
    ("MERLOT", "Red Wine"),
    ("CABERNET", "Red Wine"),
    ("PINOT NOIR", "Red Wine"),
    ("CHIANTI", "Red Wine"),
    ("SAUVIGNON", "White Wine"),
    ("CHARDONNAY", "White Wine"),
    ("RIESLING", "White Wine"),
    ("PINOT GRIGIO", "White Wine"),
    ("GAVI", "White Wine"),
    ("WINE", "Red Wine"),
    ("COFFEE", "Coffee"),
    ("ESPRESSO", "Coffee"),
    ("LATTE", "Coffee"),
    ("CAPPUCCINO", "Coffee"),
    ("AMERICANO", "Coffee"),
    ("MOCHA", "Coffee"),
    ("TEA", "Tea"),
    ("MOCKTAIL", "Mocktails"),
    ("MOJITO", "Cocktails"),
    ("MARGARITA", "Cocktails"),
    ("COCKTAIL", "Cocktails"),
    ("MIXER", "Soft Drinks"),
    ("SPRITE", "Soft Drinks"),
    ("COKE", "Soft Drinks"),
    ("COCA COLA", "Soft Drinks"),
    ("REDBULL", "Soft Drinks"),
    ("RED BULL", "Soft Drinks"),
    ("ACQUA PANNA", "Mineral Water"),
    ("SAN PELLEGRINO", "Mineral Water"),
    ("SODA WATER", "Mineral Water"),
    ("TONIC WATER", "Mineral Water"),
    ("WATER", "Mineral Water"),
    ("JUICE", "Fruit Juice & Ice Teas"),
    ("SMOOTHIE", "Fruit Juice & Ice Teas"),
    ("ICE TEA", "Fruit Juice & Ice Teas"),
    ("ICED TEA", "Fruit Juice & Ice Teas"),
]


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
        else:
            value = f"{value}".rstrip("0").rstrip(".") if "." in f"{value}" else str(value)
    text = str(value).strip()
    if not text or text.upper() == "NULL":
        return ""
    return text


def normalize_name(raw: str) -> str:
    folded = "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )
    cleaned: list[str] = []
    previous_was_space = False
    for ch in folded.strip():
        if ("A" <= ch <= "Z") or ("a" <= ch <= "z") or ch.isdigit() or ch == "-":
            cleaned.append(ch)
            previous_was_space = False
        elif ch.isspace():
            if previous_was_space or not cleaned:
                continue
            cleaned.append(" ")
            previous_was_space = True
    name = "".join(cleaned).strip()
    name = re.sub(r"\s*-\s*", "-", name)
    name = re.sub(r"\s+", " ", name).strip(" -")
    # Normalize curly apostrophes for matching (BARBERA D'ALBA)
    name = name.replace("'", "'").replace("'", "'")
    return name


def name_key(raw: str) -> str:
    return normalize_name(raw).lower()


def to_api_uom(raw: str) -> str:
    text = clean_text(raw).upper()
    if not text:
        return "pcs"
    return UOM_MAP.get(text, text.lower())


def api_json(method: str, path: str, payload: dict | None = None):
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"{API}{path}",
        data=data,
        method=method,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = resp.read().decode("utf-8")
            return resp.status, json.loads(body) if body else None
    except urllib.error.HTTPError as err:
        body = err.read().decode("utf-8", errors="replace")
        try:
            parsed = json.loads(body) if body else None
        except json.JSONDecodeError:
            parsed = {"message": body}
        return err.code, parsed


def parse_qty_uom(raw: str) -> tuple[float | None, str]:
    text = clean_text(raw)
    m = re.match(r"^\s*([0-9]+(?:\.[0-9]+)?)\s*\(([^)]+)\)\s*$", text)
    if not m:
        return None, ""
    return float(m.group(1)), m.group(2).strip().upper()


def load_recipes() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    products: list[dict] = []
    current = None
    for r in range(5, ws.max_row + 1):
        a = clean_text(ws.cell(r, 1).value)
        b = clean_text(ws.cell(r, 2).value)
        c = clean_text(ws.cell(r, 3).value)
        if not a and not b and not c:
            continue
        if not b and a:
            current = {"name": a, "ingredients": [], "subProducts": [], "row": r}
            products.append(current)
            continue
        if current is None:
            continue
        qty, uom = parse_qty_uom(c)
        if b.lower() == "ingredient":
            current["ingredients"].append(
                {"name": a, "qty": qty, "uom": uom, "row": r}
            )
        elif b.lower() == "sub product":
            current["subProducts"].append(
                {"name": a, "qty": qty, "uom": uom, "row": r}
            )
    return products


def load_menu() -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not MENU_CSV.exists():
        return out
    with MENU_CSV.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            name = clean_text(row.get("Name"))
            group = clean_text(row.get("Group"))
            try:
                rrp = float(clean_text(row.get("RRP") or "0") or 0)
            except ValueError:
                rrp = 0.0
            if not name:
                continue
            out[name_key(name)] = {
                "name": name,
                "group": group or "Specialties",
                "category": MENU_GROUP_CATEGORY.get(group, "Food"),
                "rrp": rrp,
            }
    return out


def infer_category_group(product_name: str) -> tuple[str, str]:
    padded = f" {re.sub(r'[^A-Z0-9]+', ' ', product_name.upper())} "
    for needle, group in BEVERAGE_GROUP_KEYWORDS:
        token = re.sub(r"[^A-Z0-9]+", " ", needle.upper()).strip()
        if token and f" {token} " in padded:
            return "Beverage", group
    stripped = product_name.upper().strip()
    if stripped.endswith(" BTL") or stripped.endswith(" GLS") or stripped.endswith(" TWR"):
        return "Beverage", "Liqueur"
    if stripped.endswith(" MIXER"):
        return "Beverage", "Soft Drinks"
    return "Food", "Specialties"


def empty_detail_config() -> dict:
    return {
        "altRecipeUnits": [],
        "taggedVendorProductIds": [],
        "vendorProductPrincipalQty": {},
        "vendorProductLossYield": {},
        "vendorProductComponentUom": {},
        "vendorProductLocations": {},
        "vendor": "",
        "vendorProduct": "",
        "deliveryUnitPrice": "",
        "expiryPeriodDays": "",
        "splitUse": {
            "enabled": False,
            "componentQty": "1",
            "qtyBasis": "recipe",
            "lines": [],
        },
    }


def ensure_component(
    name: str,
    by_name: dict[str, dict],
    category: str,
    group: str,
    recipe_uom: str,
) -> dict | None:
    key = name_key(name)
    existing = by_name.get(key)
    if existing:
        if not existing.get("active"):
            payload = dict(existing)
            payload["active"] = True
            payload["locationsJson"] = json.dumps(LOCATION_IDS)
            status, updated = api_json("PUT", f"/api/ingredients/{existing['id']}", payload)
            if status in (200, 201) and isinstance(updated, dict):
                by_name[key] = updated
                return updated
        return existing

    payload = {
        "id": 0,
        "companyId": COMPANY_ID,
        "componentId": "",
        "name": normalize_name(name)[:200],
        "category": category[:100],
        "group": group[:100],
        "recipeUom": recipe_uom,
        "lastPriceRecipe": 0,
        "dailyUsage": 0,
        "orderFreqDays": 0,
        "parStock": 0,
        "parStockUom": recipe_uom,
        "onHandQty": 0,
        "metricsLookbackDays": 90,
        "dailyUsageAuto": False,
        "orderFreqAuto": False,
        "storageJson": json.dumps(["Dry Store"]),
        "storageNote": "Created by Weissbrau product recipe import",
        "detailConfigJson": json.dumps(empty_detail_config()),
        "attachedProducts": 0,
        "attachedVendors": 0,
        "active": True,
        "locationsJson": json.dumps(LOCATION_IDS),
    }
    status, created = api_json("POST", "/api/ingredients", payload)
    if status not in (200, 201) or not isinstance(created, dict):
        print(f"  ! failed creating component {name!r}: {status} {created}")
        return None
    by_name[name_key(created.get("name", name))] = created
    print(f"  + component {created.get('componentId')} {created.get('name')}")
    return created


def resolve_ingredient(
    raw_name: str,
    by_name: dict[str, dict],
    line_uom: str,
) -> tuple[dict | None, str]:
    """Return (component, note)."""
    alias = INGREDIENT_ALIASES.get(raw_name.upper()) or INGREDIENT_ALIASES.get(raw_name)
    lookup = alias or raw_name
    hit = by_name.get(name_key(lookup))
    if hit:
        note = f"aliased from {raw_name}" if alias else ""
        return hit, note

    stub = STUB_COMPONENTS.get(raw_name.upper()) or STUB_COMPONENTS.get(raw_name)
    if stub:
        category, group, recipe_uom, *_rest = stub
        # Prefer recipe line UOM when present.
        if line_uom:
            recipe_uom = to_api_uom(line_uom)
        created = ensure_component(raw_name, by_name, category, group, recipe_uom)
        return created, "created stub"

    # Last resort: create under Food / Recipe Import using line UOM.
    recipe_uom = to_api_uom(line_uom) if line_uom else "g"
    created = ensure_component(
        raw_name, by_name, "Food", "Recipe Import", recipe_uom
    )
    return created, "created fallback"


def line_unit_price(component: dict, api_uom: str) -> float:
    recipe_uom = (component.get("recipeUom") or "").lower()
    if recipe_uom == api_uom.lower():
        return float(component.get("lastPriceRecipe") or 0)
    return float(component.get("lastPriceRecipe") or 0)


def build_bom_items(
    recipe: dict,
    by_name: dict[str, dict],
    report: dict,
) -> list[dict]:
    items: list[dict] = []
    for line in recipe["ingredients"]:
        qty = line.get("qty")
        if qty is None or qty <= 0:
            report["skipped_lines"].append(
                f"{recipe['name']}: bad qty for {line['name']!r} ({line.get('uom')})"
            )
            continue
        component, note = resolve_ingredient(line["name"], by_name, line.get("uom") or "")
        if not component:
            report["unmatched_ingredients"].append(
                f"{recipe['name']}: {line['name']}"
            )
            continue
        api_uom = to_api_uom(line.get("uom") or component.get("recipeUom") or "pcs")
        items.append(
            {
                "componentId": component["componentId"],
                "componentName": component.get("name") or line["name"],
                "componentUom": api_uom,
                "componentUomPrice": line_unit_price(component, api_uom),
                "quantity": qty,
            }
        )
        if note:
            report["alias_notes"].append(f"{recipe['name']}: {line['name']} → {component['name']} ({note})")
    return items


def pack_token(name: str) -> str:
    u = name.upper()
    if "(BOTTLE)" in u or re.search(r"\bBTL\b", u) or " 1 BOTTLE" in u:
        return "BTL"
    if "(GLASS)" in u or re.search(r"\bGLS\b", u):
        return "GLS"
    if "(PINT)" in u or re.search(r"\bPINT\b", u) or re.search(r"\bPNT\b", u):
        return "PINT"
    if "(TOWER)" in u or re.search(r"\bTWR\b", u):
        return "TWR"
    if "(JUG)" in u or re.search(r"\bJUG\b", u):
        return "JUG"
    if "MIXER" in u:
        return "MIXER"
    return ""


def size_ml_token(name: str) -> str:
    u = name.upper().replace("1LTR", "1000ML").replace("1 LTR", "1000ML").replace("1LT", "1000ML")
    m = re.search(r"(\d+)\s*ML\b", u)
    if m:
        return f"{int(m.group(1))}ML"
    m = re.search(r"(\d+(?:\.\d+)?)\s*L(?:TR)?\b", u)
    if m:
        return f"{int(float(m.group(1)) * 1000)}ML"
    return ""


def base_product_key(name: str) -> str:
    n = name.upper()
    for a in ["(BOTTLE)", "(GLASS)", "(PINT)", "(TOWER)", "(SHOT)", "(HOT)", "(COLD)", "(JUG)", "(SEASONAL)"]:
        n = n.replace(a, " ")
    n = re.sub(r"\b(BTL|GLS|TWR|PNT|PINT|BOTTLE|GLASS|TOWER|SHOT|MIXER|JUG)\b", " ", n)
    n = re.sub(r"\b\d+[.,]?\d*\s*(ML|LTR|L|GM|G|KG)\b", " ", n)
    n = re.sub(r"[^A-Z0-9]+", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def product_match_key(name: str) -> str:
    return "|".join([base_product_key(name), pack_token(name), size_ml_token(name)])


def find_existing_product(name: str, by_exact: dict[str, dict], by_match: dict[str, dict]) -> dict | None:
    hit = by_exact.get(name_key(name))
    if hit:
        return hit
    return by_match.get(product_match_key(name))


def product_payload(
    *,
    name: str,
    category: str,
    group: str,
    rrp: float,
    items: list[dict],
    existing: dict | None,
) -> dict:
    # Preserve variable / channel flags on update; never convert to sub-product.
    is_variable = bool(existing.get("isVariableProduct")) if existing else False
    return {
        "productId": existing.get("productId") if existing else None,
        "name": (existing.get("name") if existing else name)[:200],
        "category": category[:100],
        "group": group[:100],
        "isSubProduct": False,
        "isVariableProduct": is_variable,
        "variableMode": existing.get("variableMode") if existing else None,
        "variableChoiceQty": existing.get("variableChoiceQty") if existing else None,
        "variableOptionsJson": existing.get("variableOptionsJson") if existing else None,
        "variableMinCost": existing.get("variableMinCost") if existing else None,
        "variableMaxCost": existing.get("variableMaxCost") if existing else None,
        "b2cEnabled": True if not existing else bool(existing.get("b2cEnabled", True)),
        "b2bEnabled": False if not existing else bool(existing.get("b2bEnabled", False)),
        "b2bPackageUnit": existing.get("b2bPackageUnit") if existing else "pcs",
        "b2bSalesConfigJson": existing.get("b2bSalesConfigJson") if existing else "{}",
        "rrp": rrp,
        "posEnabled": rrp > 0,
        "active": True if not existing else bool(existing.get("active", True)),
        "companyId": COMPANY_ID,
        "locationExternalIds": LOCATION_IDS,
        "items": items,
        "packagingItems": [],
        "aliases": [],
    }


def main() -> int:
    dry_run = "--dry-run" in sys.argv
    recipes = load_recipes()
    menu = load_menu()
    print(f"Loaded {len(recipes)} products from {XLSX.name}; menu rows {len(menu)}")

    status, ingredients = api_json("GET", f"/api/ingredients?companyId={COMPANY_ID}")
    if status != 200 or not isinstance(ingredients, list):
        raise SystemExit(f"Failed to list ingredients: {status} {ingredients}")
    by_name = {name_key(i.get("name", "")): i for i in ingredients if i.get("name")}
    print(f"Components in system: {len(by_name)}")

    status, products = api_json("GET", f"/api/products?companyId={COMPANY_ID}")
    if status != 200 or not isinstance(products, list):
        raise SystemExit(f"Failed to list products: {status} {products}")
    by_product_name = {name_key(p.get("name", "")): p for p in products if p.get("name")}
    by_product_match: dict[str, dict] = {}
    for p in products:
        if not p.get("name"):
            continue
        key = product_match_key(p["name"])
        prior = by_product_match.get(key)
        if prior is None or (p.get("active") and not prior.get("active")):
            by_product_match[key] = p
    print(f"Products in system: {len(by_product_name)}")

    report: dict = {
        "created": [],
        "updated": [],
        "skipped_only_sub": [],
        "skipped_no_bom": [],
        "skipped_variable_combo": [],
        "failed": [],
        "unmatched_ingredients": [],
        "skipped_lines": [],
        "alias_notes": [],
        "deferred_sub_lines": defaultdict(list),
    }

    created = updated = skipped = failed = 0

    for recipe in recipes:
        name = recipe["name"]
        for sub in recipe["subProducts"]:
            report["deferred_sub_lines"][name].append(
                f"{sub['name']} {sub['qty']} ({sub['uom']})"
            )

        if not recipe["ingredients"] and recipe["subProducts"]:
            report["skipped_only_sub"].append(name)
            skipped += 1
            continue

        items = build_bom_items(recipe, by_name, report)
        if not items:
            report["skipped_no_bom"].append(name)
            skipped += 1
            continue

        existing = find_existing_product(name, by_product_name, by_product_match)
        # Don't overwrite combination variable products' identity; still attach BOM if allowed.
        if existing and existing.get("isVariableProduct") and (existing.get("variableMode") or "").lower() == "combination":
            # Combination products may have empty recipe by design — skip to avoid breaking packs.
            report["skipped_variable_combo"].append(name)
            skipped += 1
            continue

        menu_row = menu.get(name_key(name))
        if existing:
            category = existing.get("category") or (menu_row or {}).get("category") or "Food"
            group = existing.get("group") or (menu_row or {}).get("group") or "Specialties"
            rrp = float(existing.get("rrp") or 0)
            if rrp <= 0 and menu_row:
                rrp = float(menu_row.get("rrp") or 0)
        elif menu_row:
            category = menu_row["category"]
            group = menu_row["group"]
            rrp = float(menu_row.get("rrp") or 0)
        else:
            category, group = infer_category_group(name)
            rrp = 0.0

        # B2C requires either B2C or B2B; force B2C for new. If existing was B2B-only, keep it.
        payload = product_payload(
            name=name,
            category=category,
            group=group,
            rrp=rrp,
            items=items,
            existing=existing,
        )
        if existing and existing.get("b2bEnabled") and not existing.get("b2cEnabled"):
            payload["b2cEnabled"] = False
            payload["b2bEnabled"] = True
            payload["expiryPeriodDays"] = max(int(existing.get("expiryPeriodDays") or 0), 1)

        if dry_run:
            action = "UPDATE" if existing else "CREATE"
            print(f"[dry-run] {action} {name} ({category}/{group}) lines={len(items)} rrp={rrp}")
            continue

        if existing:
            status, result = api_json("PUT", f"/api/products/{existing['id']}", payload)
            if status in (200, 201) and isinstance(result, dict):
                updated += 1
                report["updated"].append(f"{result.get('productId')} {result.get('name')} ({len(items)} lines)")
                by_product_name[name_key(result.get("name", name))] = result
            else:
                failed += 1
                report["failed"].append(f"UPDATE {name}: {status} {result}")
                print(f"  ! update failed {name}: {status} {result}")
        else:
            status, result = api_json("POST", "/api/products", payload)
            if status in (200, 201) and isinstance(result, dict):
                created += 1
                report["created"].append(f"{result.get('productId')} {result.get('name')} ({len(items)} lines)")
                by_product_name[name_key(result.get("name", name))] = result
            else:
                failed += 1
                report["failed"].append(f"CREATE {name}: {status} {result}")
                print(f"  ! create failed {name}: {status} {result}")

        time.sleep(0.02)

    print("\n=== Summary ===")
    print(f"created={created} updated={updated} skipped={skipped} failed={failed}")
    print(f"only-sub-product recipes skipped: {len(report['skipped_only_sub'])}")
    for n in report["skipped_only_sub"]:
        print(f"  - {n}")
    print(f"no-bom skipped: {len(report['skipped_no_bom'])}")
    for n in report["skipped_no_bom"][:20]:
        print(f"  - {n}")
    if report["unmatched_ingredients"]:
        print(f"unmatched ingredient lines: {len(report['unmatched_ingredients'])}")
        for line in report["unmatched_ingredients"][:30]:
            print(f"  - {line}")
    if report["failed"]:
        print("failures:")
        for line in report["failed"][:40]:
            print(f"  - {line}")

    out_path = ROOT / "data" / "weissbrau-product-recipe-import-report.json"
    serializable = {
        **report,
        "deferred_sub_lines": dict(report["deferred_sub_lines"]),
        "counts": {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "failed": failed,
        },
    }
    out_path.write_text(json.dumps(serializable, indent=2), encoding="utf-8")
    print(f"Wrote {out_path}")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
